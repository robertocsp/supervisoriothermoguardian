import datetime
from django.contrib import admin

from . import models
from django.http import HttpResponse, HttpResponseRedirect
from openpyxl import Workbook

@admin.register(models.Cliente)
class ClienteAdmin(admin.ModelAdmin):
    list_display = ('nome', 'descricao')



@admin.register(models.Datalog)
class DatalogAdmin(admin.ModelAdmin):
    list_display = ('id', 'valor', 'parametro_unidade', 'datahora','parametro', 'parametro_nome',)
    list_filter = (
        ('parametro__modulo__circuito__nome'),
        ('parametro__modulo__no_slave'),
        ('parametro__nome')
    )

    def parametro_nome(self, instance):
        return instance.parametro.nome

    def parametro_unidade(self, instance ):
        return instance.parametro.unidade

@admin.register(models.Logerros)
class LogerrosAdmin(admin.ModelAdmin):
    list_display = ('cod', 'descricao', 'datahora')

@admin.register(models.Superaquecimentoconfig)
class SuperaquecimentoconfigAdmin(admin.ModelAdmin):
    list_display = ('nome', 'succao_no_slave', 'succao_no_parametro', 'evaporacao_no_slave', 'evaporacao_no_parametro', 'min_superaquecimento', 'max_superaquecimento')

@admin.register(models.Superaquecimentolog)
class SuperaquecimentologAdmin(admin.ModelAdmin):
    list_display = ('resultado', 'superaquecimento', 'superaquecimentoconfig', 'datahora')

    # list_per_page = sys.maxsize
    list_per_page = 500
    actions = ['export_excel']

    def export_excel(self, request, queryset):
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename={date}-superaquecimento.xlsx'.format(
            date=datetime.datetime.now().strftime('%Y-%m-%d'),
        )
        workbook = Workbook()

        # Get active worksheet/tab
        worksheet = workbook.active
        worksheet.title = 'Superaquecimento'

        # Define the titles for columns
        columns = [
            'Resultado',
            'Superaquecimento',
            'DataHora',

        ]
        row_num = 1

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        # Iterate through all movies
        for item in queryset:
            row_num += 1

            # Define the data for each cell in the row
            row = [
                item.resultado,
                'Não' if item.superaquecimento else 'Sim',
                item.datahora,

            ]

            # Assign the data for each cell of the row
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        workbook.save(response)

        return response

    export_excel.short_description = "Exportar para Excel itens selecionados"



admin.site.register(models.Circuito)
admin.site.register(models.Modulo)
admin.site.register(models.Parametro)

@admin.register(models.Alarme)
class AlarmeAdmin(admin.ModelAdmin):
    list_display = ('nome', 'no_slave', 'no_parametro', 'minimo', 'maximo', 'em_alarme', 'habilitado', 'ultimo_email_enviado', 'ultimo_sms_enviado')

@admin.register(models.Alarmelog)
class AlarmelogAdmin(admin.ModelAdmin):
    list_display = ('resultado', 'alarme', 'datahora', 'email_enviado', 'sms_enviado', 'parametro_nome', 'registro_datalog')

    def parametro_nome(self, instance):
        return instance.datalog.parametro.nome

    def registro_datalog(self, instance):
        return instance.datalog.datahora



